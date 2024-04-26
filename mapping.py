from openpyxl import Workbook
import os

directory = "D:\\834 File\\Current 837"
excel_file = "D:\\Test\\MappingSheet.xlsx"

wb = Workbook()
sheet = wb.active
#
if sheet.cell(row=1,column=1).value is None:
    sheet.cell(row=1,column=1).value = "Scenario ID"
    sheet.cell(row=1,column=2).value = "Member ID"

filecount = 0

for filename in os.listdir(directory):
    if filename.endswith('.837'):
        filecount += 1

print(f"Found {filecount} EDI 837 Files in directory")
row_num = 2
for filename in os.listdir(directory):
    file_path = os.path.join(directory, filename)
    with open(file_path, 'r') as file:
        file_content = file.read()
        # print(file_content)

        lines = file_content.splitlines()

        for line in lines:
            if line.startswith('CLM'):
                parts = line.split('*')
                Scenario_number = parts[1]
                sheet.cell(row=row_num, column=1).value = Scenario_number

            if line.startswith('NM1*IL'):
                parts = line.split('*')
                Member_number = parts[-1]
                tilde_index = Member_number.find('~')
                updated_number = Member_number[:tilde_index]
                # modified_number = str(updated_number)
                # MemberID = modified_number[:-2] + '-' + modified_number[-2:]
                sheet.cell(row=row_num, column=2).value = updated_number
        row_num += 1;

wb.save(excel_file)
wb.close()









