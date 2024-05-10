import os
from openpyxl import Workbook

# Set Source and Destination paths
Source_path = "D:\\Test\\Current 837"
Dest_path = "D:\\Test\\Member_Details837.xlsx"

# Create a new Excel workbook
workbook = Workbook()
sheet = workbook.active

# Write headers to Excel Sheet if it is new workbook
if sheet.cell(1,1).value is None:
    sheet.cell(1,1,"CLM Segment")
    sheet.cell(1,2,"Member Number")
    sheet.cell(1, 3, "Member FirstName")
    sheet.cell(1, 4, "Member lastName")
    sheet.cell(1, 5, "DOB")
    sheet.cell(1, 6, "Address")

# Initialize row count
row_num = 2
nmi_line_index = None

# Iterate through all file in source folder and extract details
for filename in os.listdir(Source_path):
    if filename.endswith(".837"):
        file_path = os.path.join(Source_path,filename)
        print(f"Processing file:{filename}")

        with open(file_path, 'r') as file:
            lines = file.read()
        multi_line = '\n'.join(lines.split('~'))
        with open(file_path, 'w') as file:
            file.writelines(multi_line)

        # Read  the content of the file
        with open(file_path,"r") as file:
            lines = file.readlines()

        nmi_line = ""
        clm_line = ""
        dmg_line = ""
        n3_line = ""

        # Loop through each line in file
        for index,line in enumerate(lines):
            if "NM1*IL*" in line:
                nmi_line = line
                nmi_line_index = index
            elif "CLM" in line:
                clm_line = line
            elif "DMG*D8" in line:
                dmg_line = line
            elif "N3" in line and nmi_line_index is not  None and index==nmi_line_index+1:
                n3_line = line

        # Extract line
        clm_Segment = clm_line.split("*")[1] if clm_line else ""
        member_number = nmi_line.split("*")[9][:10] if nmi_line else ""
        member_fn = nmi_line.split("*")[4]
        member_ln = nmi_line.split("*")[3]
        member_dob = dmg_line.split("*")[2]
        member_addr = n3_line.split("*")[1]

        # Write extracted data to excel
        sheet.cell(row_num, 1, clm_Segment)
        sheet.cell(row_num, 2, member_number)
        sheet.cell(row_num, 3, member_fn)
        sheet.cell(row_num, 4, member_ln)
        sheet.cell(row_num, 5, member_dob)
        sheet.cell(row_num, 6, member_addr)

        # After replacement add ~ at end of each line
        with open(file_path, 'r') as file:
            lines = file.readlines()
        lines_with_tilde = [line.rstrip('\n') + '~' for line in lines]
        with open(file_path, 'w') as file:
            file.writelines(lines_with_tilde)

        # convert multiple lines into single line
        with open(file_path, 'r') as file:
            lines = file.readlines()
        single_line = ' '.join(line.strip() for line in lines)
        with open(file_path, 'w') as file:
            file.writelines(single_line)

        # Move to next row
        row_num += 1

# Save to Excel workbook
workbook.save(Dest_path)
workbook.close()
print(f"Data Extraction Completed")