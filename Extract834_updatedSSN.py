import os
from openpyxl import Workbook

# Define input file and output file
input_file = "D:\\Test\\834 File"
output_file = "D:\\Test\\EDI834.xlsx"

# Initialize Excel Workbook
wb = Workbook()
ws = wb.active

intRow = 1
filecount = 0

for filename in os.listdir(input_file):
    if filename.lower().endswith('.edi'):
        filecount += 1

print(f"Found {filecount} EDI Files in directory")

def extract_edi_data(file_path, ws, intRow):
    file_content = ""
    LastName = ""
    FirstName = ""
    ID = ""
    DOB = ""
    intStartIndex = 0
    intEndIndex = 0
    strSegment = ""

    # Read EDI data from file
    with open(file_path,'r') as edi_file:
        file_content = edi_file.read()

    # Find Segment Start with NM1*IL*1*
    intStartIndex = file_content.find("NM1*IL*1*")

    # Process each segment Found
    while intStartIndex != -1:
        # Find End of Segment
        intEndIndex = file_content.find("~",intStartIndex)
        if intEndIndex == -1:
            break

        # Extract Segment
        strSegment = file_content[intStartIndex:intEndIndex]

        # Split Segment into parts
        arrSegmentParts = strSegment.split("*")

        # Extract LastName, FirstName and ID
        if len(arrSegmentParts) >= 4:
            LastName = arrSegmentParts[3]
        if len(arrSegmentParts) >= 5:
            FirstName = arrSegmentParts[4]
        if len(arrSegmentParts) >= 9:
            ID = arrSegmentParts[9]  # Extract last 9 Digit


        # Write To excel
        ws.cell(row=intRow, column=1).value = LastName
        ws.cell(row=intRow,column=2).value = FirstName
        ws.cell(row=intRow, column=3).number_format = '@'
        ws.cell(row=intRow, column=3).value = ID

        # Move to the next Row in Excel
        intRow += 1

        # Find next Segment Start with NM1*IL*1*
        intStartIndex = file_content.find("NM1*IL*1*",intEndIndex)


for filename in os.listdir(input_file):
    if filename.lower().endswith('.edi'):
        file_path = os.path.join(input_file,filename)
        print(f"Processing file:{filename}")
        extract_edi_data(file_path,ws,intRow)
        intRow = ws.max_row+1

# Save Excel Workbook
wb.save(output_file)

print("Data Extraction Completed")




