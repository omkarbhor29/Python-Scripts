import os
from openpyxl import load_workbook

src_folder = r"D:\\Test"

excel_path = r"D:\\Test\\Mastersheet.xlsx"

wb = load_workbook(excel_path)
ws = wb.active

# Iterate each row in excel with filename
for row in ws.iter_rows(min_row=2,values_only=True):
    file_name = str(row[0]).strip()
    member_number = str(row[1]).strip()
    file_path = os.path.join(src_folder,file_name)

    print("Processing File",file_name)

    # split the lines end with ~
    with open(file_path, 'r') as file:
        lines = file.read()
    multi_line = '\n'.join(lines.split('~'))
    with open(file_path, 'w') as file:
        file.writelines(multi_line)

    # Replace member number
    if os.path.isfile(file_path):
        with open(file_path,'r') as file:
            file_content = file.read()
            lines = file_content.splitlines()

            replaced = False
            for i,line in enumerate(lines):
                if "NM1*IL" in line:
                    lines[i] = line[:-10] + member_number
                    replaced = True
                    break

            if replaced:
                updated_content = "\n".join(lines)
                with open(file_path,'w') as updated_file:
                    updated_file.write(updated_content)
                print("Member replaced sucessfully in file:",file_path)
            else:
                print("File Not found",file_path)

    # After replacement add ~ at end of each line
    with open(file_path, 'r') as file:
        lines = file.readlines()
    lines_with_tilde = [line.rstrip('\n') + '~' for line in lines]
    with open(file_path, 'w') as file:
        file.writelines(lines_with_tilde)

    # convert multiple lines into single line
    with open(file_path, 'r') as file:
        lines = file.readlines()
    single_line = ' '.join(line.strip() for line in lines)
    with open(file_path, 'w') as file:
        file.writelines(single_line)

wb.close()
print("Processing Complete")


