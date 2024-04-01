import openpyxl

# Define input file and output file
input_file = r"D:\\Test\\EDI834.xlsx"
output_file = input_file

Excel = openpyxl.load_workbook(input_file)
inputsheet = Excel.worksheets[0]  # First sheet
outputSheet = Excel.create_sheet(index=1)  # Create new Sheet for output

# Loop through the column in the input sheet
for col in range(1,inputsheet.max_column+1):
    concat_value = ""

    # Loop through the rows in the column
    for row in range(1,inputsheet.max_row+1):
        concat_value += f"'{inputsheet.cell(row=row,column=col).value}',"
        print(concat_value)

    # Remove the trailing comma
    concat_value = concat_value[:-1]

    # write concatenated value to the output sheet in single row
    outputSheet.cell(row=1, column=col).value = concat_value

# Save the Workbook
Excel.save(output_file)

# Close the Workbook
Excel.close()