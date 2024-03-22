import os
from openpyxl import load_workbook

src_folder = r"D:\\Test\\Processed 837"

excel_path = r"D:\\Test\\Mastersheet.xlsx"

wb = load_workbook(excel_path)
ws = wb.active

for row in ws.iter_rows(min_row=2,values_only=True):
    file_name = str(row[0]).strip()
    member_number = str(row[1]).strip()
    file_path = os.path.join(src_folder,file_name)

    print("Processing File",file_name)

    if os.path.isfile(file_path):
        with open(file_path,'r') as file:
            file_content = file.read()
            lines = file_content.splitlines()

            replaced = False
            for i,line in enumerate(lines):
                if "NM1*IL" in line:
                    lines[i] = line[:-11] + member_number
                    replaced = True
                    break

            if replaced:
                updated_content = "\n".join(lines)
                with open(file_path,'w') as updated_file:
                    updated_file.write(updated_content)
                print("Member replaced sucessfully in file:",file_path)
            else:
                print("File Not found",file_path)

wb.close()
print("Processing Complete")


